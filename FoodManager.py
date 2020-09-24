import openpyxl
import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.mime.base import MIMEBase
import smtplib

path_output = r'C:\Users\nekry\Desktop\PYTHON PROJECTS\Foodmanager\output\\'

def send_msg(post, text):
	msg = MIMEMultipart() #создание объекта сообщения
	message = "Заказ: \n" + text #сообщение которое отправялется
	password = "kingking2281337" # пароль от почты с которой идет отправление
	msg['From'] = "testtesttesttest9990@gmail.com" #почта с которой отправляется
	msg['To'] = post # почта куда отправляется
	msg['Subject'] = "Поставки" # тема сообщения

	msg.attach(MIMEText(message, 'plain'))

	server = smtplib.SMTP('smtp.gmail.com: 587') # подключение к серверу отправки сообщений

	server.starttls() # хз точно, но как понял делает подключение к почте безопасным

	server.login(msg['From'], password) #логин в почту с которой будет отправляться сообщение

	server.sendmail(msg['From'], msg['To'], msg.as_string()) # отправка сообщения
	server.quit()
	print ("successfully sent email to %s:" % (msg['To'])) # вывод того что сообщение отправилось


path_manage = r'C:\Users\nekry\Desktop\PYTHON PROJECTS\Foodmanager\some\manage.xlsx'#пусть к таблице с основной информацией
wb = openpyxl.load_workbook(filename=path_manage) #открытие файла для чтения с основной информацией
wb.active = 0 #установка рабочего листа на первый
sheet = wb.active

post_dict = {} #словарь с ключами - почта поставщики, значения - продукты
post_list = [] #список только с почтой поставщиков 
for i in range(2, 4):
	post_dict[sheet['C'+str(i)].value] = sheet['D'+str(i)].value #формирование словаря
wb.save(path_manage) #закрытие файла из которго читали
post_list = [key for key in post_dict]# формирование списка поставщиков


path_prod = r'C:\Users\nekry\Desktop\PYTHON PROJECTS\Foodmanager' #в какой папке искать файл с отчетом
files = [i for i in os.listdir(path_prod) if i.endswith('.xlsx')]#выбираются только файлы с расщирение xlsx

new = max(files, key=os.path.getctime)#самый новый файл среди них
files = [os.path.join(path_prod, file) for file in files] #добавление полного пути к файлам
new_path = max(files, key=os.path.getctime) #самый новый файл с путем к нему

wb = openpyxl.load_workbook(filename=new_path) #открытие файла для чтения с продуктами
wb.active = 0 #установка рабочего листа на первый
sheet = wb.active

for post in post_list:# цикл по поставщикам продуктов
	wb = openpyxl.load_workbook(filename=path_manage) #открытие файла для чтения с основной информацией
	date = datetime.now()
	i = 2 # переменная для пробега по всем строкам тамблицы
	text = '' # список продуктов
	while sheet['A'+str(i)].value:
		if sheet['A' + str(i)].value in post_dict.get(post):#если взятые продукт относится к рассматриваемому поставщику то работаем
			text = text + sheet['A' + str(i)].value + ' - ' + str(sheet['B' + str(i)].value) + ' ' + sheet['C' + str(i)].value + '\n' # добавление в текст сообщения продукта количества и единиц измерения
		i += 1
	send_msg(post, text) # ф-ия отправки сообщения
wb.save(new) # закрытие файла с продуктами
